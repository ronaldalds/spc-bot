import os
import time
import openpyxl
from process.spc import include
from dotenv import load_dotenv
from pyrogram.types import Message
from pyrogram import Client
import concurrent.futures
from driver.formatador import formatar_data

load_dotenv()

running = False

def handle_start_include_spc(client: Client, message: Message):
    global running
    if not running:
        running = True
        # Verifique se a mensagem contém um documento e se o tipo MIME do documento é "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if message.document and message.document.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            # Criando Pool
            limite_threads = 4
            # Baixe o arquivo XLSX
            message.reply_text("Preparando arquivo XLSX")
            file_path = message.download()
            # Processar o arquivo XLSX conforme necessário
            try:
                try:
                    workbook = openpyxl.load_workbook(file_path)
                except openpyxl.utils.exceptions.InvalidFileException:
                    message.reply_text("O arquivo fornecido não é um arquivo XLSX válido.")
                    return
                
                sheet = workbook.active
                max_row = sheet.max_row
                lista = []
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if row[headers.index('Tipo de Pessoa')] == 'física':
                        try:
                            cpf_cnpj = str(row[headers.index('CPF/CNPJ')])
                            data_nascimento = formatar_data(str(row[headers.index('Data Nascimento')]))
                            ddd = str(row[headers.index('DDD')])
                            celular = str(row[headers.index('Celular')])
                            cep = str(row[headers.index('CEP')])
                            logradouro = str(row[headers.index('Logradouro')])
                            numero = str(row[headers.index('Número')])
                            complemento = str(row[headers.index('Complemento')])
                            bairro = str(row[headers.index('Bairro')])
                            data_vencimento = formatar_data(str(row[headers.index('Data Vencimento')]))
                            data_compra = formatar_data(str(row[headers.index('Data Compra')]))
                            cod_cliente = str(row[headers.index('Cod Cliente')])
                            valor_debito = str(row[headers.index('Valor do Débito')]).replace('.', ',')

                            # Se chegou até aqui, os dados são válidos, então adiciona à lista
                            lista.append((
                                cpf_cnpj,
                                data_nascimento,
                                ddd,
                                celular,
                                cep,
                                logradouro,
                                numero,
                                complemento,
                                bairro,
                                data_vencimento,
                                data_compra,
                                cod_cliente,
                                valor_debito
                                ))
                        except Exception as e:
                            print(f"Error: na linha {len(lista) + 1}, {e}")

                message.reply_text(f"Processando arquivo XLSX com {len(lista)} clientes...")
                def executar(arg):
                    if running:
                        try:
                            include(
                                cpf_cnpj = arg[0],
                                data_nascimento = arg[1],
                                ddd = arg[2],
                                celular = arg[3],
                                cep = arg[4],
                                logradouro = arg[5],
                                numero = arg[6],
                                complemento = arg[7],
                                bairro = arg[8],
                                data_vencimento = arg[9],
                                data_compra = arg[10],
                                cod_cliente = arg[11],
                                valor_debito = arg[12],
                                )
                        except Exception as e:
                            print(f"Error executing: cpf:{arg[0]} - {e}")
                    else:
                        message.reply_text(f"Inclusão SPC cpf:{arg[0]} parado.")
                with concurrent.futures.ThreadPoolExecutor(max_workers=limite_threads) as executor:
                    executor.map(executar, lista)
                
            # ...
            finally:
            # Excluir o arquivo XLSX
                time.sleep(1)
                os.remove(file_path)
            # Responder à mensagem do usuário com o resultado do processamento do arquivo
            message.reply_text("O arquivo XLSX foi processado com sucesso!")
            running = False
        else:
            # Responder à mensagem do usuário com uma mensagem de erro
            message.reply_text("Por favor, envie um arquivo XLSX para processar.")
    else:
        message.reply_text("Inclusão SPC em execução.")

def handle_stop_include_spc(client: Client, message: Message):
    global running
    running = False
    message.reply_text("Pedido de parada iniciado...")

def handle_status_include_spc(client: Client, message: Message):
    global running
    try:
        if running:
            message.reply_text("Inclusão SPC em execução.")
        else:
            message.reply_text("Inclusão SPC parado.")
    except:
        message.reply_text("Inclusão SPC parado.")