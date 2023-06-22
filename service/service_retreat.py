import openpyxl
import os
import time
from process.recolhimento import recolhimento
from dotenv import load_dotenv
from pyrogram.types import Message
from pyrogram import Client
import concurrent.futures
from driver.formatador import formatar_int, formatar_documento

load_dotenv()

running = False

lojas_mk1 = {
    "LOJA ALCÂNTARAS": 2,
    "LOJA ANAPURUS": 2,
    "LOJA ARARENDÁ": 2,
    "LOJA BOA VIAGEM": 3,
    "LOJA CAMOCIM": 10,
    "LOJA CARIRÉ": 2,
    "LOJA CARNAUBAL": 2,
    "LOJA GUARACIABA DO NORTE": 5,
    "LOJA HIDROLÂNDIA": 2,
    "LOJA IBIAPINA": 2,
    "LOJA IPUEIRAS": 2,
    "LOJA ITATIRA": 2,
    "LOJA LISIEUX": 2,
    "LOJA MATA ROMA": 2,
    "LOJA MERUOCA": 2,
    "LOJA MONSENHOR TABOSA": 2,
    "LOJA NOVO ORIENTE": 3,
    "LOJA PEDRO II": 3,
    "LOJA PIRACURUCA": 3,
    "LOJA PIRIPIRI": 4,
    "LOJA RERIUTABA": 2,
    "LOJA SANTA QUITÉRIA": 3,
    "LOJA SÃO BERNARDO": 2,
    "LOJA TAMBORIL": 2,
    "LOJA UBAJARA": 2,
    "LOJA VARJOTA": 4,
}

lojas_mk3 = {
    "LOJA CASTANHAL": 8,
    "LOJA VIGIA": 5,
    "LOJA TERRA ALTA": 4,
    "LOJA ICOARACI": 5,
    "LOJA MARITUBA": 9,
    "LOJA VILA DOS CABANOS": 8,
    "LOJA BARCARENA": 4,
    "LOJA MAGUARI": 4,
    "LOJA ABAETETUBA": 12,
    "LOJA TUCURUI": 5,
    "LOJA TAILÂNDIA": 4,
    "LOJA MOJU": 4,
    "LOJA MOCAJUBA": 3,
    "LOJA BAIÃO": 2
}

def __gerar_lista(workbook):
    sheet = workbook.active
    max_row = sheet.max_row
    headers = [cell.value for cell in sheet[1]]
    lista = []

    for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
        if (str(row[headers.index("Qtd Conexoes")]) == "1") and (row[headers.index("OS Cancelamento ou Recolhimento")] == "N"):
            try:
                mk = formatar_int(row[headers.index("MK")])
                contrato = formatar_int(row[headers.index("Contrato")])
                conexao_associada = formatar_int(row[headers.index("Conexao Associada")])
                cpf = formatar_documento(row[headers.index("Documento/Codigo")])['cpf']
                cod = formatar_documento(row[headers.index("Documento/Codigo")])['cod']
                tipo_da_os = row[headers.index("Tipo OS")]
                grupo_atendimento_os = str(row[headers.index("Grupo Atendimento OS")]).strip()
                detalhe_os = row[headers.index("Detalhe OS")]
                loja = str(row[headers.index("Loja")])

                lista.append((
                    mk,
                    contrato,
                    conexao_associada,
                    cpf,
                    cod,
                    tipo_da_os,
                    grupo_atendimento_os,
                    detalhe_os,
                    loja,
                    ))
            except Exception as e:
                print(f"Error: na linha {len(lista) + 1}, {e}")

    return lista

def __limpa_lista(lista):
    # cria dict com todas as lojas que tem limite de O.S
    dicionario = dict(lojas_mk1, **lojas_mk3)
    lista_resultante = []
    # cria um dict com todos os valores 0 para contar as ocorrencias
    ocorrencias_dict = {chave: 0 for chave in dicionario}
    
    # percorre a lista de O.S de recolhimento que existe
    for item in lista:
        chave = item[8] # posicao da tupla onde esta o que voce quer contar a ocorrencia
        if chave in dicionario:
            valor_limite = dicionario[chave] # quantidade de O.S que pdoe ser abertas por loja
            if ocorrencias_dict[chave] < valor_limite:
                lista_resultante.append(item)
                ocorrencias_dict[chave] += 1
        else:
            # continua na lista as lojas que nao tiverem cadastrado um limite de O.S
            lista_resultante.append(item)

    return lista_resultante

def handle_start_retreat_mk(client: Client, message: Message):
    global running
    if not running:
        running = True
        # Verifique se a mensagem contém um documento e se o tipo MIME do documento é "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if message.document and message.document.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            # Quantidade de itens na Pool
            limite_threads = 20

            # Baixe o arquivo XLSX
            file_path = message.download()
            message.reply_text("Preparando arquivo XLSX")
            
            # Processar o arquivo XLSX conforme necessário
            try:
                try:
                    workbook = openpyxl.load_workbook(file_path)
                except openpyxl.utils.exceptions.InvalidFileException:
                    message.reply_text("O arquivo fornecido não é um arquivo XLSX válido.")
                    return
                
                # lista com todas as os
                lista = __gerar_lista(workbook)
                message.reply_text(f"Processando arquivo XLSX com {len(lista)} O.S...")

                # lista com base na quantidade maxima de os que pode ser abertas por loja
                lista = __limpa_lista(lista)
                message.reply_text(f"Arquivo XLSX ajustado para {len(lista)} O.S...")
                def executar(arg):
                    if running:
                        try:
                            recolhimento(
                                mk=arg[0],
                                contrato=arg[1],
                                conexao_associada=arg[2],
                                cpf=arg[3],
                                cod=arg[4],
                                tipo_da_os=arg[5],
                                grupo_atendimento_os=arg[6],
                                detalhe_os=arg[7],
                                loja=arg[8]
                            )
                        except Exception as e:
                            print(f"Error executing na função executar:mk:{arg[0]} contrato:{arg[1]} conexão:{arg[2]} {e}")
                    else:
                        message.reply_text(f"Recolhimento mk:{arg[0]} contrato:{arg[1]} conexão:{arg[2]} parado.")

                # Criando Pool
                with concurrent.futures.ThreadPoolExecutor(max_workers=limite_threads) as executor:
                    executor.map(executar, lista)
            # ...
            finally:
            # Excluir o arquivo XLSX
                time.sleep(1)
                os.remove(file_path)
            # Responder à mensagem do usuário com o resultado do processamento do arquivo
            message.reply_text("O arquivo XLSX foi processado com sucesso!")
            running = False
        else:
            # Responder à mensagem do usuário com uma mensagem de erro
            message.reply_text("Por favor, envie um arquivo XLSX para processar.")
    else:
        message.reply_text("Recolhimento em execução.")
    
def handle_stop_retreat_mk(client: Client, message: Message):
    global running
    running = False
    message.reply_text("Pedido de parada iniciado...")

def handle_status_retreat_mk(client: Client, message: Message):
    global running
    try:
        if running:
            message.reply_text("Recolhimento em execução")
        else:
            message.reply_text("Recolhimento parado")
    except:
        message.reply_text("Recolhimento parado")