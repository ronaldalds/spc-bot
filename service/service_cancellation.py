import openpyxl
import os
import time
from process.cancelamento import cancelamento
from dotenv import load_dotenv
from pyrogram.types import Message
from pyrogram import Client
import concurrent.futures
from driver.formatador import formatar_data, formatar_incidencia, formatar_valor_multa, formatar_int

load_dotenv()

running = False

def handle_start_cancellation_mk(client: Client, message: Message):
    global running
    if not running:
        running = True
        # Verifique se a mensagem contém um documento e se o tipo MIME do documento é "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if message.document and message.document.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            # Quantidade de itens na Pool
            limite_threads = 20

            # Baixe o arquivo XLSX
            file_path = message.download()
            message.reply_text("Preparando arquivo XLSX")
            
            # Processar o arquivo XLSX conforme necessário
            try:
                try:
                    workbook = openpyxl.load_workbook(file_path)
                except openpyxl.utils.exceptions.InvalidFileException:
                    message.reply_text("O arquivo fornecido não é um arquivo XLSX válido.")
                    return

                sheet = workbook.active
                max_row = sheet.max_row
                lista = []
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    try:
                        mk = formatar_int(row[headers.index("MK")])
                        cod_pessoa = formatar_int(row[headers.index("Cod Pessoa")])
                        contrato = formatar_int(row[headers.index("Contrato")])
                        detalhes_cancelamento = row[headers.index("Detalhes Cancelamento")]
                        tipo_da_os = row[headers.index("Tipo OS")]
                        grupo_atendimento_os = str(row[headers.index("Grupo Atendimento OS")]).strip()
                        relato_do_problema = row[headers.index("Relato do problema")]
                        incidencia_multa = formatar_incidencia(row[headers.index("Incidencia de Multa")])
                        valor_multa = formatar_valor_multa(row[headers.index("Valor Multa")])
                        vencimento_multa = formatar_data(row[headers.index("Data Vcto Multa Contratual")])
                        planos_contas = row[headers.index("Planos de Contas")]

                        # Se chegou até aqui, os dados são válidos, então adiciona à lista
                        lista.append((
                            mk,
                            cod_pessoa,
                            contrato,
                            detalhes_cancelamento,
                            tipo_da_os,
                            grupo_atendimento_os,
                            relato_do_problema,
                            incidencia_multa,
                            valor_multa,
                            vencimento_multa,
                            planos_contas
                            ))
                    except Exception as e:
                        print(f"Error: na linha {len(lista) + 1}, {e}")

                message.reply_text(f"Processando arquivo XLSX com {len(lista)} contratos...")
                def executar(arg):

                    if running:
                        try:
                            cancelamento(
                                mk=arg[0],
                                cod_pessoa=arg[1],
                                contrato=arg[2],
                                detalhes_cancelamento=arg[3],
                                tipo_da_os=arg[4],
                                grupo_atendimento_os=arg[5],
                                relato_do_problema=arg[6],
                                incidencia_multa=arg[7],
                                valor_multa=arg[8],
                                vencimento_multa=arg[9],
                                planos_contas=arg[10]
                            )
                        except Exception as e:
                            print(f"Error executing na função executar:mk:{arg[0]} cod:{arg[1]} contrato:{arg[2]} {e}")
                    else:
                        message.reply_text(f"Cancelamento mk:{arg[0]} cod:{arg[1]} contrato:{arg[2]} parado.")

                # Criando Pool
                with concurrent.futures.ThreadPoolExecutor(max_workers=limite_threads) as executor:
                    executor.map(executar, lista)
            # ...
            finally:
            # Excluir o arquivo XLSX
                time.sleep(1)
                os.remove(file_path)
            # Responder à mensagem do usuário com o resultado do processamento do arquivo
            message.reply_text("O arquivo XLSX foi processado com sucesso!")
            running = False
        else:
            # Responder à mensagem do usuário com uma mensagem de erro
            message.reply_text("Por favor, envie um arquivo XLSX para processar.")
    else:
        message.reply_text("Cancelamento em execução.")
    
def handle_stop_cancellation_mk(client: Client, message: Message):
    global running
    running = False
    message.reply_text("Pedido de parada iniciado...")

def handle_status_cancellation_mk(client: Client, message: Message):
    global running
    try:
        if running:
            message.reply_text("Cancelamento em execução")
        else:
            message.reply_text("Cancelamento parado")
    except:
        message.reply_text("Cancelamento parado")